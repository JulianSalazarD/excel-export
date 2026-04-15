"""
insert_cotizacion.py
Inserta un DatosCotizacion en el xlsx de control de cotizaciones Melectra.

Uso:
    python insert_cotizacion.py <archivo.docx> [archivo2.docx ...]

El script extrae los datos de cada .docx y los inserta en el xlsx como
una nueva fila. Si el número de cotización ya existe en la hoja lo omite.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from extract_cotizacion import CotizacionExtractor
from models import DatosCotizacion


# ---------------------------------------------------------------------------
# Configuración
# ---------------------------------------------------------------------------

XLSX_PATH   = Path("docs/COTIZACIONES 2026. - copia.xlsx")
HEADER_ROW  = 5          # fila donde están los encabezados (1-indexed)
COL_MEDIO   = 2          # B  — MEDIO POR EL CUAL SE DIO CUENTA
COL_NUMERO  = 3          # C  — N° COTIZACIÓN
COL_EMPRESA = 4          # D  — NOMBRE EMPRESA
COL_NOMBRE  = 5          # E  — ENCARGADO-SOLICITANTE
COL_SERVICIO= 6          # F  — SERVICIO
COL_CORREO  = 7          # G  — CORREO
COL_TELEFONO= 8          # H  — TELEFONO
COL_VALOR   = 9          # I  — VALOR
COL_ESTADO  = 10         # J  — ESTADO
COL_FACTURA = 13         # M  — N° FACTURA (fecha de la cotización)

DEFAULT_MEDIO  = ""
DEFAULT_ESTADO = "RECIBIDA"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_valor(raw: Optional[str]) -> Optional[float]:
    """Convierte '$9,800,000' | '$ 9.800.000' | '1.700.000' a float."""
    from extract_cotizacion import _parse_raw_valor
    return _parse_raw_valor(raw) if raw else None


def _find_sheet(wb: openpyxl.Workbook) -> Worksheet:
    """
    Devuelve la hoja de datos activa.
    Prioridad: hoja activa del libro → primera hoja que no sea 'DESPLEGABLES'.
    """
    for name in wb.sheetnames:
        if "DESPLEGABLE" not in name.upper():
            return wb[name]
    return wb.active


def _existing_numeros(ws: Worksheet) -> set[str]:
    """Retorna el conjunto de números de cotización ya presentes."""
    numeros: set[str] = set()
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        val = row[COL_NUMERO - 1]
        if val:
            numeros.add(str(val).strip())
    return numeros


# ---------------------------------------------------------------------------
# Inserción principal
# ---------------------------------------------------------------------------

def insert_row(ws: Worksheet, datos: DatosCotizacion) -> None:
    """Agrega una fila al final de los datos existentes en la hoja."""
    # Encontrar la primera fila vacía después del encabezado
    next_row = HEADER_ROW + 1
    for row in ws.iter_rows(min_row=HEADER_ROW + 1):
        if all(cell.value is None for cell in row):
            break
        next_row += 1

    ws.cell(row=next_row, column=COL_MEDIO,    value=DEFAULT_MEDIO)
    ws.cell(row=next_row, column=COL_NUMERO,   value=datos.numero)
    ws.cell(row=next_row, column=COL_EMPRESA,  value=datos.empresa)
    ws.cell(row=next_row, column=COL_NOMBRE,   value=datos.nombre)
    ws.cell(row=next_row, column=COL_SERVICIO, value=datos.servicio)
    ws.cell(row=next_row, column=COL_CORREO,   value=datos.correo)
    ws.cell(row=next_row, column=COL_TELEFONO, value=datos.telefono)
    ws.cell(row=next_row, column=COL_VALOR,    value=_parse_valor(datos.valor_total))
    ws.cell(row=next_row, column=COL_ESTADO,   value=DEFAULT_ESTADO)
    ws.cell(row=next_row, column=COL_FACTURA,  value=datos.fecha)


def insert_cotizacion(
    datos: DatosCotizacion,
    xlsx_path: Path = XLSX_PATH,
    skip_duplicates: bool = True,
) -> bool:
    """
    Inserta `datos` en el xlsx.

    Retorna True si se insertó, False si se omitió por duplicado.
    """
    wb = load_workbook(xlsx_path)
    ws = _find_sheet(wb)

    if skip_duplicates and datos.numero:
        existentes = _existing_numeros(ws)
        if datos.numero in existentes:
            print(f"  [OMITIDO] {datos.numero} ya existe en la hoja.")
            return False

    insert_row(ws, datos)
    wb.save(xlsx_path)
    return True


