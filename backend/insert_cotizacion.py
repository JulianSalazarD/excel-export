"""
insert_cotizacion.py
Inserta un DatosCotizacion en el xlsx de control de cotizaciones Melectra.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from extract_cotizacion import _parse_raw_valor
from models import DatosCotizacion
from xlsx_manager import COL_MAP, create_backup, find_data_sheet, find_header_row

XLSX_PATH      = Path("docs/COTIZACIONES 2026. - copia.xlsx")
DEFAULT_ESTADO = "RECIBIDA"


def _parse_valor(raw: Optional[str]) -> Optional[float]:
    return _parse_raw_valor(raw) if raw else None


def _existing_pairs(ws: Worksheet, data_start: int) -> set[tuple[str, str]]:
    col_numero = COL_MAP["numero"] - 1
    col_correo = COL_MAP["correo"] - 1
    pairs: set[tuple[str, str]] = set()
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        numero = row[col_numero]
        correo = row[col_correo]
        if numero and correo:
            pairs.add((str(numero).strip(), str(correo).strip().lower()))
    return pairs


def insert_row(ws: Worksheet, datos: DatosCotizacion, data_start: int) -> None:
    """Inserta la fila en la posición que mantiene el orden por número de cotización."""
    new_numero = str(datos.numero).strip() if datos.numero else None

    if new_numero:
        # Buscar posición ordenada ascendentemente por número de cotización
        target_row = data_start
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            existing_numero = row[COL_MAP["numero"] - 1]
            if existing_numero is None:
                break  # fin de datos
            if str(existing_numero).strip() > new_numero:
                break  # posición de inserción encontrada
            target_row += 1
    else:
        # Sin número → agregar al final (comportamiento anterior)
        target_row = data_start
        for row in ws.iter_rows(min_row=data_start):
            if all(cell.value is None for cell in row):
                break
            target_row += 1

    # Insertar fila en la posición encontrada (desplaza filas existentes hacia abajo)
    ws.insert_rows(target_row)

    ws.cell(row=target_row, column=COL_MAP["medio"],    value=datos.medio or "")
    ws.cell(row=target_row, column=COL_MAP["numero"],   value=datos.numero)
    ws.cell(row=target_row, column=COL_MAP["empresa"],  value=datos.empresa)
    ws.cell(row=target_row, column=COL_MAP["nombre"],   value=datos.nombre)
    ws.cell(row=target_row, column=COL_MAP["servicio"], value=datos.servicio)
    ws.cell(row=target_row, column=COL_MAP["correo"],   value=datos.correo)
    ws.cell(row=target_row, column=COL_MAP["telefono"], value=datos.telefono)
    ws.cell(row=target_row, column=COL_MAP["valor_total"], value=_parse_valor(datos.valor_total))
    ws.cell(row=target_row, column=COL_MAP["estado"],   value=datos.estado or DEFAULT_ESTADO)
    ws.cell(row=target_row, column=COL_MAP["trabajo_realizado_en"], value=datos.trabajo_realizado_en or "")
    ws.cell(row=target_row, column=COL_MAP["orden_servicio"],       value=datos.orden_servicio or "")
    ws.cell(row=target_row, column=COL_MAP["observacion"], value=datos.observacion or "")


def _add_dropdown_validations(ws: Worksheet, data_start: int) -> None:
    """Crea los dropdowns de medio (col B) y estado (col J).

    openpyxl elimina las validaciones por extensión al cargar el archivo
    (warning: 'Data Validation extension is not supported and will be removed').
    Esta función las recrea desde la hoja DESPLEGABLES antes de guardar.
    """
    from openpyxl.worksheet.datavalidation import DataValidation, DataValidationList

    if ws.data_validations is None:
        ws.data_validations = DataValidationList()
    else:
        ws.data_validations.dataValidation.clear()

    # Encontrar última fila con datos
    last_row = data_start
    for row in ws.iter_rows(min_row=data_start):
        if all(c.value is None for c in row):
            break
        last_row = row[0].row

    # Medio (col B) → DESPLEGABLES!$B$2:$B$9
    dv_medio = DataValidation(type="list", formula1="=DESPLEGABLES!$B$2:$B$9", allow_blank=True)
    dv_medio.sqref = f"$B${data_start}:$B${last_row}"
    ws.data_validations.dataValidation.append(dv_medio)

    # Estado (col J) → DESPLEGABLES!$D$2:$D$6
    dv_estado = DataValidation(type="list", formula1="=DESPLEGABLES!$D$2:$D$6", allow_blank=True)
    dv_estado.sqref = f"$J${data_start}:$J${last_row}"
    ws.data_validations.dataValidation.append(dv_estado)


def insert_cotizacion(
    datos: DatosCotizacion,
    xlsx_path: Path = XLSX_PATH,
    skip_duplicates: bool = True,
    sheet_name: Optional[str] = None,
) -> bool:
    wb = load_workbook(xlsx_path)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = find_data_sheet(wb)
    header_row = find_header_row(ws)
    data_start = header_row + 1

    if skip_duplicates and datos.numero and datos.correo:
        key = (datos.numero.strip(), datos.correo.strip().lower())
        if key in _existing_pairs(ws, data_start):
            return False

    insert_row(ws, datos, data_start)
    _add_dropdown_validations(ws, data_start)
    create_backup(xlsx_path)
    wb.save(xlsx_path)
    return True
