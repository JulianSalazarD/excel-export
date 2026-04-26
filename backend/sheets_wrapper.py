"""
Wrapper para leer hojas de un archivo Excel.

Uso:
    python sheets_wrapper.py <xlsx_path>
    # Salida: JSON con lista de hojas
"""

import json
import sys
from pathlib import Path

# Añadir el directorio padre al path para importar los módulos
sys.path.insert(0, str(Path(__file__).parent.parent))

from openpyxl import load_workbook


def main() -> None:
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Uso: sheets_wrapper.py <xlsx_path>"}))
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])

    if not xlsx_path.exists():
        print(json.dumps({"error": f"Archivo no encontrado: {xlsx_path}"}))
        sys.exit(1)

    try:
        wb = load_workbook(xlsx_path)
        sheets = wb.sheetnames
        wb.close()

        print(json.dumps({"sheets": sheets}))
        sys.exit(0)

    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
