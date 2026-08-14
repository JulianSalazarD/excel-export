"""
xlsx_manager.py
Carga, edita y guarda el libro de cotizaciones Melectra.

Usa Polars para el manejo en memoria y openpyxl para leer/escribir
conservando estilos y fórmulas en el resto del libro.
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
import polars as pl
from openpyxl.worksheet.worksheet import Worksheet

from models import DatosCotizacion

# Nombres de meses en español para detectar la hoja del mes actual
_MESES_ES = [
    "ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
    "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE",
]

# ---------------------------------------------------------------------------
# Mapeo campo → columna (1-indexed)
# ---------------------------------------------------------------------------

COL_MAP: dict[str, int] = {
    "medio":                2,   # B — MEDIO POR EL CUAL SE DIO CUENTA
    "numero":               3,   # C — N° COTIZACIÓN
    "empresa":              4,   # D — NOMBRE EMPRESA
    "nombre":               5,   # E — ENCARGADO-SOLICITANTE
    "servicio":             6,   # F — SERVICIO
    "correo":               7,   # G — CORREO
    "telefono":             8,   # H — TELEFONO
    "valor_total":          9,   # I — VALOR
    "estado":               10,  # J — ESTADO
    "trabajo_realizado_en": 11,  # K — TRABAJO REALIZADO EN
    "orden_servicio":       12,  # L — ORDEN DE SERVICIO MELECTRA
    "numero_factura":       13,  # M — N° FACTURA
    "observacion":          14,  # N — OBSERVACIÓN (fecha extraída + notas)
}

CAMPOS = list(COL_MAP.keys())

MAX_BACKUPS = 3


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def find_data_sheet(wb: openpyxl.Workbook) -> Worksheet:
    for name in wb.sheetnames:
        if "DESPLEGABLE" not in name.upper():
            return wb[name]
    return wb.active


def find_header_row(ws: Worksheet) -> int:
    """Detecta dinámicamente la fila de encabezados buscando 'MEDIO'."""
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            if cell.value and "MEDIO" in str(cell.value).upper():
                return cell.row
    return 5  # fallback


def list_sheets(xlsx_path: Path) -> list[str]:
    """Retorna los nombres de hojas del libro, excluyendo DESPLEGABLE."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    return [n for n in wb.sheetnames if "DESPLEGABLE" not in n.upper()]


def find_month_sheet(sheets: list[str]) -> Optional[str]:
    """Busca la hoja que coincida con el mes actual (ej: 'MAYO 2026')."""
    mes_actual = _MESES_ES[datetime.now().month - 1]
    anio_actual = str(datetime.now().year)
    # Primero buscar coincidencia exacta mes+año
    for s in sheets:
        if mes_actual in s.upper() and anio_actual in s:
            return s
    # Luego solo mes
    for s in sheets:
        if mes_actual in s.upper():
            return s
    return sheets[0] if sheets else None


def _cell_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------

def create_backup(xlsx_path: Path) -> Path:
    """
    Copia el xlsx en <xlsx_dir>/backups/<stem>_<timestamp>.xlsx.
    Guardarlo junto al archivo fuente garantiza una ruta escribible y fácil
    de encontrar para el usuario, tanto en dev como dentro de un AppImage.
    Conserva solo los MAX_BACKUPS más recientes por cada archivo.
    """
    backup_dir = xlsx_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = backup_dir / f"{xlsx_path.stem}_{ts}{xlsx_path.suffix}"
    shutil.copy2(xlsx_path, dest)

    pattern = f"{xlsx_path.stem}_*{xlsx_path.suffix}"
    backups = sorted(backup_dir.glob(pattern))
    for old in backups[:-MAX_BACKUPS]:
        old.unlink(missing_ok=True)

    return dest


# ---------------------------------------------------------------------------
# Carga
# ---------------------------------------------------------------------------

def load_filas(xlsx_path: Path) -> pl.DataFrame:
    """
    Lee las filas de datos del xlsx y retorna un DataFrame de Polars.
    Todos los campos son Utf8 (o null). Incluye '_row' con el nro. de fila.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = find_data_sheet(wb)
    header_row = find_header_row(ws)
    data_start = header_row + 1

    records: list[dict] = []
    for row in ws.iter_rows(min_row=data_start, values_only=False):
        vals = {f: _cell_str(row[col - 1].value) for f, col in COL_MAP.items()}
        if not any(v for v in vals.values()):
            continue
        vals["_row"] = row[0].row
        records.append(vals)

    schema = {f: pl.Utf8 for f in CAMPOS}
    schema["_row"] = pl.Int32

    if not records:
        return pl.DataFrame(schema=schema)

    return pl.DataFrame(records, infer_schema_length=None).cast(schema)


# ---------------------------------------------------------------------------
# Creación de hoja desde plantilla
# ---------------------------------------------------------------------------

HEADER_TEXTS: dict[str, str] = {
    "medio":                "MEDIO POR EL CUAL SE DIO CUENTA",
    "numero":               "N° COTIZACIÓN",
    "empresa":              "NOMBRE EMPRESA",
    "nombre":               "ENCARGADO- SOLICITANTE",
    "servicio":             "SERVICIO",
    "correo":               "CORREO",
    "telefono":             "TELEFONO",
    "valor_total":          "VALOR",
    "estado":               "ESTADO",
    "trabajo_realizado_en": "TRABAJO REALIZADO  EN",
    "orden_servicio":       "ORDEN DE SERVICIO MELECTRA",
    "numero_factura":       "N° \nFACTURA",
    "observacion":          "OBSERVACIÓN",
}

HEADER_ROW = 5


def create_template_sheet(xlsx_path: Path, sheet_name: str) -> str:
    """Crea una hoja nueva con la plantilla de encabezados y validaciones.

    Returns el nombre de la hoja creada (puede tener sufijo si ya existe).
    """
    wb = openpyxl.load_workbook(xlsx_path)

    # Si ya existe una hoja con ese nombre, agregar sufijo
    original = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        counter += 1
        sheet_name = f"{original} ({counter})"

    ws = wb.create_sheet(title=sheet_name)

    from openpyxl.styles import Alignment, Border, Font, Side, PatternFill

    # --- Filas superiores (título e info) ---
    ws.row_dimensions[1].height = 3.75
    ws.row_dimensions[2].height = 58.5
    ws.row_dimensions[3].height = 14.25
    ws.row_dimensions[4].height = 9.0

    # Título principal (E2:N2 fusionado)
    ws.merge_cells("E2:N2")
    title_cell = ws["E2"]
    title_cell.value = "CONTROL DE COTIZACIONES"
    title_cell.font = Font(name="Calibri", size=20, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # B2:D2 fusionado (vacío)
    ws.merge_cells("B2:D2")

    # Fila 3: CÓDIGO, VERSIÓN, VIGENCIA, Página
    title_font = Font(name="Arial", size=10, bold=True)
    title_align = Alignment(horizontal="center")
    for col, text in [(2, "CÓDIGO"), (5, "VERSIÓN"), (9, "VIGENCIA"), (12, "Página")]:
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = title_font
        cell.alignment = title_align

    # Fusiones de fila 3
    ws.merge_cells("G3:H3")
    ws.merge_cells("J3:K3")
    ws.merge_cells("M3:N3")

    # --- Encabezados en fila 5 ---
    header_font = Font(name="Arial", size=10, bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(style="thin")
    header_border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    for name, col_idx in sorted(COL_MAP.items(), key=lambda x: x[1]):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=HEADER_TEXTS.get(name, ""))
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    # --- Ancho de columnas ---
    col_widths = {
        "A": 1.0, "C": 18.4, "D": 21.7, "E": 25.1, "F": 33.9,
        "G": 30.0, "H": 12.0, "I": 14.0, "K": 14.4, "L": 11.0, "N": 45.0,
    }
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # --- Alto de fila de encabezados ---
    ws.row_dimensions[HEADER_ROW].height = 53.25

    # --- Validaciones dropdown (Medio y Estado) ---
    from openpyxl.worksheet.datavalidation import DataValidation

    dv_medio = DataValidation(
        type="list",
        formula1="=DESPLEGABLES!$B$2:$B$9",
        allow_blank=True,
    )
    dv_medio.error = "Seleccione un medio válido de la lista"
    dv_medio.errorTitle = "Medio inválido"
    ws.add_data_validation(dv_medio)
    dv_medio.sqref = f"B{HEADER_ROW + 1}:B{HEADER_ROW + 100}"

    dv_estado = DataValidation(
        type="list",
        formula1="=DESPLEGABLES!$D$2:$D$6",
        allow_blank=True,
    )
    dv_estado.error = "Seleccione un estado válido de la lista"
    dv_estado.errorTitle = "Estado inválido"
    ws.add_data_validation(dv_estado)
    dv_estado.sqref = f"J{HEADER_ROW + 1}:J{HEADER_ROW + 100}"

    wb.save(xlsx_path)
    return sheet_name


# ---------------------------------------------------------------------------
# Guardado
# ---------------------------------------------------------------------------

def save_filas(xlsx_path: Path, filas: list[dict]) -> Path:
    """
    Escribe la lista de dicts al xlsx preservando estilos.
    Crea backup antes de guardar y retorna la ruta del backup.
    """
    backup_path = create_backup(xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = find_data_sheet(wb)
    header_row = find_header_row(ws)
    data_start = header_row + 1

    # Borrar filas de datos existentes
    if ws.max_row >= data_start:
        ws.delete_rows(data_start, ws.max_row - data_start + 1)

    # Escribir nuevas filas
    for i, fila in enumerate(filas):
        row_idx = data_start + i
        for campo, col in COL_MAP.items():
            val = fila.get(campo)
            ws.cell(row=row_idx, column=col, value=val if val else None)

    wb.save(xlsx_path)
    return backup_path


# ---------------------------------------------------------------------------
# Conversión DatosCotizacion ↔ dict
# ---------------------------------------------------------------------------

def datos_to_dict(datos: DatosCotizacion) -> dict:
    return {f: getattr(datos, f, None) for f in CAMPOS}


def dict_to_datos(d: dict) -> DatosCotizacion:
    return DatosCotizacion(**{f: d.get(f) for f in CAMPOS})
